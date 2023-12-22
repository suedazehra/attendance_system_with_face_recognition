import numpy as np
import cv2
import face_recognition
import os
import datetime as dt
import pandas as pd
from openpyxl import *

Path = '/Users/suedazehra/Desktop/projects/Attendance-with-face-recognition/Images'
images = []
classNames = []
imgList = os.listdir(Path)

# Numara bilgilerini içeren bir sözlük
numbers_dict = {
    'BRITNEY SPEARS': '20973384',
    'ANDY SAMBERG': '210239238',
    'ELON MUSK': '928723246',
    'BILL GATES': '2032934432'
    # Diğer kişileri buraya ekleyin
}

for img in imgList:
    curImg = cv2.imread(f'{Path}/{img}')
    if curImg is not None:
        images.append(curImg)
        classNames.append(os.path.splitext(img)[0])

def FindEncoding(images):
    encodings = []
    for img in images:
        if img is not None:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encode = face_recognition.face_encodings(img)[0]
            encodings.append(encode)
    return encodings

def MarkAttendance(name):
    file_path = '/Users/suedazehra/Desktop/projects/Attendance-with-face-recognition/Attendance.xlsx'
    date = dt.date.today()
    time_now = dt.datetime.now().strftime('%H:%M:%S')

    # Kişinin adına göre numarasını al
    number = numbers_dict.get(name, 'Bilgi Yok')

    if not os.path.isfile(file_path):
        # Dosya yoksa yeni bir Excel dosyası oluştur
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Number', 'Date', 'Time'])

        # İlk veriyi ekle
        sheet.append([name, number, str(date), time_now])

        # Dosyayı kaydet
        workbook.save(file_path)
    else:
        # Dosya varsa mevcut Excel dosyasını yükle
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Aynı kişinin daha önce eklenip eklenmediğini kontrol et
        existing_names = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row)]
        if name not in existing_names:
            # Dosyaya yeni veriyi ekle
            sheet.append([name, number, str(date), time_now])

            # Dosyayı kaydet
            workbook.save(file_path)


cap = cv2.VideoCapture(0)
encodeKnownList = FindEncoding(images)

while True:
    try:
        _, image = cap.read()
        imageS = cv2.resize(image, (0, 0), None, 0.25, 0.25)
        if imageS is not None:
            imageS = cv2.cvtColor(imageS, cv2.COLOR_BGR2RGB)
            faces = face_recognition.face_locations(imageS)
            curEncodings = face_recognition.face_encodings(imageS, faces)

            # encodeKnownList = FindEncoding(images)

            for encodeFace, FaceLoc in zip(curEncodings, faces):
                matches = face_recognition.compare_faces(encodeKnownList, encodeFace)
                faceDist = face_recognition.face_distance(encodeKnownList, encodeFace)
                matchIndex = np.argmin(faceDist)

                if matches[matchIndex]:
                    name = classNames[matchIndex].upper()  
                    y1, x2, y2, x1 = FaceLoc
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    cv2.rectangle(image, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(image, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                    cv2.putText(image, name, (x1 + 6, y2 - 10), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

                    MarkAttendance(name)
                    
    except Exception as e:
        print(f'Hata: {e}')

    cv2.imshow("Attendance System", image)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
